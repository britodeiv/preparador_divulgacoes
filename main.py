
from __future__ import annotations

import json
import os
import re
import threading
import traceback
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tkinter import END, StringVar, filedialog, messagebox

import customtkinter as ctk

try:
    import pythoncom
    import win32com.client
except Exception:
    pythoncom = None
    win32com = None

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
except Exception:
    Workbook = None
    load_workbook = None


APP_NAME = "Preparador de Divulgações"
APP_VERSION = "1.2.0"
DEVELOPER = "Desenvolvido por Deivid Brito"
DEFAULT_SPECIAL_BLOCK = ""
DEFAULT_BATCH = 580

EMAIL_PATTERN = re.compile(
    r"(?<![A-Z0-9._%+\-])([A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,})(?![A-Z0-9._%+\-])",
    re.I,
)
SUBJECT_PREFIXES = re.compile(r"^\s*((ENC|RE|RES|FW|FWD)\s*:\s*)+", re.I)
BAD_FILENAME = re.compile(r'[<>:"/\\|?*\x00-\x1F]+')
MULTI_SPACE = re.compile(r"\s+")
MULTI_DOT = re.compile(r"\.{2,}")


def app_data_dir() -> Path:
    root = Path(os.getenv("APPDATA", Path.home()))
    folder = root / "PreparadorDivulgacoes"
    folder.mkdir(parents=True, exist_ok=True)
    return folder


CONFIG_FILE = app_data_dir() / "config.json"
LOG_FILE = app_data_dir() / "preparador.log"


def log_error(text: str) -> None:
    try:
        with LOG_FILE.open("a", encoding="utf-8") as file:
            file.write(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {text}\n")
    except Exception:
        pass


def save_json(path: Path, data: dict) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def load_json(path: Path, default: dict) -> dict:
    try:
        if path.exists():
            return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        pass
    return default


def clean_subject(text: str) -> str:
    subject = str(text or "").strip()
    old = None
    while old != subject:
        old = subject
        subject = SUBJECT_PREFIXES.sub("", subject).strip()
    return subject or "Divulgação"


def safe_folder_name(text: str) -> str:
    value = BAD_FILENAME.sub(" ", str(text))
    value = MULTI_SPACE.sub(" ", value).strip(" .")
    return value[:120] or f"Divulgação {datetime.now():%Y-%m-%d}"


def remove_accents(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def normalize_candidate(raw: object) -> str:
    text = str(raw or "")
    text = text.replace("\u00a0", " ").replace("\t", " ")
    text = text.strip().lower()
    text = remove_accents(text)

    # Resíduos frequentes do Word, navegador e planilhas.
    text = text.replace("mailto:", "")
    text = text.replace("https://", "")
    text = text.replace("http://", "")
    text = text.replace("www.", "")
    text = text.replace('"', "")
    text = text.replace("'", "")
    text = text.replace("<", "")
    text = text.replace(">", "")
    text = text.replace("\\", "")
    text = text.replace("/", "")
    text = text.replace(";", " ")
    text = text.replace(",", " ")
    text = text.replace(":", " ")
    text = text.replace("%20", "")
    text = text.replace("%", "")
    text = text.replace("&", "")
    text = MULTI_SPACE.sub(" ", text).strip()
    return text


def split_possible_emails(raw: object) -> list[str]:
    text = normalize_candidate(raw)
    if not text:
        return []

    # Extrai e-mails válidos que já possam estar no meio do texto.
    found = EMAIL_PATTERN.findall(text)
    if found:
        return [email.lower().strip() for email in found]

    # Mantém candidatos com @ para relatório de correção/inválidos.
    pieces = re.split(r"[\s;,\n\r]+", text)
    return [piece.strip() for piece in pieces if piece.strip() and "@" in piece]


def correct_email(candidate: str) -> tuple[str | None, str]:
    email = candidate.strip().lower()

    email = email.replace("@.", "@")
    email = email.replace(".@", "@")
    email = MULTI_DOT.sub(".", email)
    email = email.strip(" .,:;<>\"'")

    # Correções seguras e explicitamente usadas pela equipe.
    if email.endswith("@gmail.com.br"):
        email = email[:-3]
    if email.endswith("@gmail"):
        email += ".com"
    elif email.endswith("@hotmail"):
        email += ".com"
    elif email.endswith("@outlook"):
        email += ".com"
    elif email.endswith("@yahoo"):
        email += ".com.br"

    if email.count("@") != 1:
        return None, "Sem @ ou com mais de um @"

    local, domain = email.split("@", 1)
    if not local or not domain:
        return None, "Parte local ou domínio vazio"

    if email.startswith(".") or local.endswith("."):
        return None, "Ponto em posição inválida"

    if domain.startswith(".") or domain.endswith("."):
        return None, "Domínio terminando ou começando com ponto"

    if domain.endswith(("hotmail", "gmail", "outlook", "yahoo", "b")):
        return None, "Domínio incompleto"

    if "@googlegroups" in email:
        return None, "Google Groups não permitido"

    if "." not in domain:
        return None, "Domínio sem extensão"

    if not EMAIL_PATTERN.fullmatch(email):
        return None, "Formato inválido"

    return email, "OK"


@dataclass
class ProcessResult:
    raw_cells: int
    found_candidates: int
    valid_before_dedup: int
    duplicates: list[str]
    invalid: list[tuple[str, str]]
    blocked: list[str]
    debate_removed: list[str]
    final: list[str]
    block_sheet: str


class ExcelReader:
    @staticmethod
    def ensure() -> None:
        if load_workbook is None:
            raise RuntimeError("A biblioteca de Excel não foi incluída corretamente.")

    @classmethod
    def read_todos(cls, path: Path) -> tuple[list[str], list[tuple[str, str]], int, int]:
        cls.ensure()
        workbook = load_workbook(path, read_only=True, data_only=True)
        valid: list[str] = []
        invalid: list[tuple[str, str]] = []
        raw_cells = 0
        candidates_count = 0

        try:
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for value in row:
                        if value is None:
                            continue
                        raw_cells += 1
                        candidates = split_possible_emails(value)
                        candidates_count += len(candidates)
                        for candidate in candidates:
                            corrected, reason = correct_email(candidate)
                            if corrected:
                                valid.append(corrected)
                            else:
                                invalid.append((candidate, reason))
        finally:
            workbook.close()

        return valid, invalid, raw_cells, candidates_count

    @classmethod
    def read_blocklist(cls, path: Path) -> tuple[set[str], str]:
        cls.ensure()
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            selected = None
            for name in workbook.sheetnames:
                if name.strip().lower() == "resumo":
                    selected = name
                    break

            if selected is None:
                raise ValueError(
                    "A aba 'Resumo' não foi encontrada na planilha de e-mails que voltaram.\n\n"
                    "Selecione a planilha correta, pois esta versão usa somente a aba Resumo."
                )

            blocked: set[str] = set()
            sheet = workbook[selected]

            for row in sheet.iter_rows(values_only=True):
                for value in row:
                    for candidate in split_possible_emails(value):
                        corrected, _ = correct_email(candidate)
                        if corrected:
                            blocked.add(corrected)

            if not blocked:
                raise ValueError(
                    "A aba 'Resumo' foi encontrada, mas não contém e-mails reconhecíveis."
                )

            return blocked, selected
        finally:
            workbook.close()



class OutlookBridge:
    def __init__(self) -> None:
        if pythoncom is None or win32com is None:
            raise RuntimeError("A integração com o Outlook não foi incluída no executável.")
        pythoncom.CoInitialize()
        try:
            self.app = win32com.client.Dispatch("Outlook.Application")
        except Exception as exc:
            raise RuntimeError(
                "Não foi possível acessar o Outlook clássico.\n\n"
                "Abra o Outlook (clássico), confirme que a conta está configurada e tente novamente."
            ) from exc

    def selected_mail(self):
        explorer = self.app.ActiveExplorer()
        if explorer is not None and explorer.Selection.Count >= 1:
            item = explorer.Selection.Item(1)
            if str(getattr(item, "MessageClass", "")).startswith("IPM.Note"):
                return item

        inspector = self.app.ActiveInspector()
        if inspector is not None:
            item = inspector.CurrentItem
            if str(getattr(item, "MessageClass", "")).startswith("IPM.Note"):
                return item

        raise RuntimeError(
            "Nenhum e-mail foi selecionado.\n\n"
            "No Outlook clássico, clique no e-mail recebido ou deixe a mensagem aberta."
        )

    def selected_info(self) -> dict:
        item = self.selected_mail()
        return {
            "subject": clean_subject(getattr(item, "Subject", "")),
            "sender": str(getattr(item, "SenderName", "")),
            "received": str(getattr(item, "ReceivedTime", "")),
        }

    def create_draft(self, recipients: list[str]) -> str:
        original = self.selected_mail()
        draft = original.Forward()
        subject = clean_subject(getattr(original, "Subject", ""))

        draft.Subject = subject
        draft.To = ""
        draft.CC = ""
        draft.BCC = "; ".join(recipients)

        # Remove o cabeçalho de encaminhamento, mantendo conteúdo, imagens e anexos.
        try:
            draft.HTMLBody = original.HTMLBody
        except Exception:
            draft.Body = original.Body

        draft.Display()
        return subject


class App(ctk.CTk):
    def __init__(self) -> None:
        super().__init__()
        self.title(f"{APP_NAME} v{APP_VERSION}")
        self.geometry("1120x790")
        self.minsize(980, 680)

        ctk.set_appearance_mode("System")
        ctk.set_default_color_theme("green")

        config = load_json(CONFIG_FILE, {})

        self.todos_file: Path | None = None
        self.block_file: Path | None = None
        self.output_folder: Path | None = None
        self.result: ProcessResult | None = None
        self.batches: list[list[str]] = []
        self.current_batch = 0
        self.outlook_subject = ""

        self.type_var = StringVar(value="Campanha especial")
        self.todos_var = StringVar(value="Nenhum arquivo .Todos selecionado")
        self.block_var = StringVar(value=config.get("block_file", "Nenhuma planilha selecionada"))
        self.output_var = StringVar(value=config.get("output_folder", str(Path.home() / "Documents")))
        self.batch_var = StringVar(value=str(config.get("batch_size", DEFAULT_BATCH)))
        self.special_block_var = StringVar(value=config.get("special_block_email", DEFAULT_SPECIAL_BLOCK))
        self.status_var = StringVar(value="Selecione a planilha .Todos para começar.")
        self.outlook_var = StringVar(value="Nenhum e-mail carregado do Outlook.")

        if Path(self.block_var.get()).exists():
            self.block_file = Path(self.block_var.get())
        if Path(self.output_var.get()).exists():
            self.output_folder = Path(self.output_var.get())

        self._build_ui()

    def _build_ui(self) -> None:
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        header = ctk.CTkFrame(self, corner_radius=0)
        header.grid(row=0, column=0, sticky="ew")
        ctk.CTkLabel(
            header,
            text="PREPARADOR DE DIVULGAÇÕES",
            font=ctk.CTkFont(size=25, weight="bold"),
        ).pack(anchor="w", padx=22, pady=(16, 2))
        ctk.CTkLabel(
            header,
            text="Limpeza da .Todos, bloqueios, lotes em CCO e Outlook clássico.",
        ).pack(anchor="w", padx=22, pady=(0, 3))
        ctk.CTkLabel(
            header,
            text=DEVELOPER,
            font=ctk.CTkFont(size=12, slant="italic"),
        ).pack(anchor="w", padx=22, pady=(0, 14))

        tabs = ctk.CTkTabview(self)
        tabs.grid(row=1, column=0, padx=18, pady=14, sticky="nsew")
        for tab_name in ("1. Limpeza", "2. Resultado", "3. Outlook"):
            tabs.add(tab_name)

        self._build_clean_tab(tabs.tab("1. Limpeza"))
        self._build_result_tab(tabs.tab("2. Resultado"))
        self._build_outlook_tab(tabs.tab("3. Outlook"))

        footer = ctk.CTkFrame(self, corner_radius=0)
        footer.grid(row=2, column=0, sticky="ew")
        footer.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(footer, textvariable=self.status_var, anchor="w").grid(
            row=0, column=0, padx=18, pady=9, sticky="ew"
        )
        ctk.CTkLabel(footer, text=DEVELOPER).grid(
            row=0, column=1, padx=18, pady=9
        )

    def _build_clean_tab(self, tab) -> None:
        tab.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(tab, text="Tipo da divulgação", font=ctk.CTkFont(weight="bold")).grid(
            row=0, column=0, padx=16, pady=(18, 8), sticky="w"
        )
        type_frame = ctk.CTkFrame(tab, fg_color="transparent")
        type_frame.grid(row=0, column=1, padx=16, pady=(18, 8), sticky="w")
        ctk.CTkRadioButton(
            type_frame, text="Campanha especial",
            variable=self.type_var, value="Campanha especial"
        ).pack(side="left", padx=(0, 20))
        ctk.CTkRadioButton(
            type_frame, text="Campanha comum",
            variable=self.type_var, value="Campanha comum"
        ).pack(side="left")

        ctk.CTkLabel(
            tab,
            text="E-mail opcional a excluir em campanha especial",
            font=ctk.CTkFont(weight="bold"),
        ).grid(row=1, column=0, padx=16, pady=8, sticky="w")
        special_frame = ctk.CTkFrame(tab)
        special_frame.grid(row=1, column=1, padx=16, pady=8, sticky="ew")
        special_frame.grid_columnconfigure(0, weight=1)
        ctk.CTkEntry(
            special_frame,
            textvariable=self.special_block_var,
            placeholder_text="exemplo@organizacao.org",
        ).grid(row=0, column=0, padx=10, pady=10, sticky="ew")

        self._file_row(tab, 2, "Planilha da campanha (.Todos)", "Selecionar .Todos", self.select_todos, self.todos_var)
        self._file_row(tab, 3, "E-mails que voltaram / pediram para sair", "Selecionar bloqueados", self.select_block, self.block_var)
        self._file_row(tab, 4, "Pasta para salvar o resultado", "Selecionar pasta", self.select_output, self.output_var)

        options = ctk.CTkFrame(tab)
        options.grid(row=5, column=0, columnspan=2, padx=16, pady=15, sticky="ew")
        ctk.CTkLabel(options, text="E-mails por lote").pack(side="left", padx=(14, 8), pady=14)
        ctk.CTkEntry(options, textvariable=self.batch_var, width=90).pack(side="left", padx=8, pady=14)
        ctk.CTkLabel(options, text="Padrão: 580 | Máximo: 600 | Sempre em CCO").pack(
            side="left", padx=12, pady=14
        )

        ctk.CTkButton(
            tab,
            text="LIMPAR, COMPARAR E GERAR LISTA FINAL",
            command=self.process_async,
            height=50,
            font=ctk.CTkFont(size=16, weight="bold"),
        ).grid(row=6, column=0, columnspan=2, padx=16, pady=14, sticky="ew")

        self.busy = ctk.CTkProgressBar(tab, mode="indeterminate")
        self.busy.grid(row=7, column=0, columnspan=2, padx=16, pady=8, sticky="ew")
        self.busy.stop()

        rules_text = (
            "A ferramenta converte para minúsculas, remove espaços, tabulações, aspas, "
            "http, www, mailto, acentos e resíduos; corrige casos simples, elimina inválidos, "
            "remove repetidos, compara exclusivamente com a aba Resumo da lista de bloqueados "
            "e aplica uma exclusão opcional configurável para campanhas especiais."
        )
        ctk.CTkLabel(tab, text=rules_text, wraplength=940, justify="left").grid(
            row=7, column=0, columnspan=2, padx=16, pady=12, sticky="w"
        )

    def _file_row(self, tab, row, title, button_text, command, variable) -> None:
        ctk.CTkLabel(tab, text=title, font=ctk.CTkFont(weight="bold")).grid(
            row=row, column=0, padx=16, pady=14, sticky="w"
        )
        frame = ctk.CTkFrame(tab)
        frame.grid(row=row, column=1, padx=16, pady=8, sticky="ew")
        frame.grid_columnconfigure(1, weight=1)
        ctk.CTkButton(frame, text=button_text, command=command, width=175).grid(
            row=0, column=0, padx=10, pady=10
        )
        ctk.CTkLabel(frame, textvariable=variable, anchor="w", wraplength=650).grid(
            row=0, column=1, padx=10, pady=10, sticky="ew"
        )

    def _build_result_tab(self, tab) -> None:
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(2, weight=1)

        self.summary_label = ctk.CTkLabel(
            tab,
            text="Nenhum processamento realizado.",
            font=ctk.CTkFont(size=18, weight="bold"),
            anchor="w",
            justify="left",
        )
        self.summary_label.grid(row=0, column=0, padx=16, pady=16, sticky="ew")

        actions = ctk.CTkFrame(tab)
        actions.grid(row=1, column=0, padx=16, pady=8, sticky="ew")
        actions.grid_columnconfigure((0, 1, 2), weight=1)
        ctk.CTkButton(actions, text="Abrir pasta", command=self.open_folder).grid(
            row=0, column=0, padx=8, pady=10, sticky="ew"
        )
        ctk.CTkButton(actions, text="Gerar relatório novamente", command=self.export_again).grid(
            row=0, column=1, padx=8, pady=10, sticky="ew"
        )
        ctk.CTkButton(actions, text="Copiar resumo", command=self.copy_summary).grid(
            row=0, column=2, padx=8, pady=10, sticky="ew"
        )

        self.result_log = ctk.CTkTextbox(tab)
        self.result_log.grid(row=2, column=0, padx=16, pady=16, sticky="nsew")
        self._result_log("Aguardando processamento.")

    def _build_outlook_tab(self, tab) -> None:
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(5, weight=1)

        ctk.CTkLabel(
            tab,
            text="No Outlook clássico, selecione o e-mail recebido ou deixe a mensagem aberta.",
            font=ctk.CTkFont(size=16, weight="bold"),
        ).grid(row=0, column=0, padx=16, pady=(20, 8), sticky="w")

        ctk.CTkButton(
            tab, text="CARREGAR E-MAIL DO OUTLOOK",
            command=self.load_outlook, height=45
        ).grid(row=1, column=0, padx=16, pady=8, sticky="ew")

        ctk.CTkLabel(tab, textvariable=self.outlook_var, wraplength=930, justify="left").grid(
            row=2, column=0, padx=16, pady=10, sticky="w"
        )

        self.outlook_summary = ctk.CTkLabel(
            tab,
            text="Processe a lista antes de abrir os lotes.",
            font=ctk.CTkFont(size=17, weight="bold"),
            anchor="w",
            justify="left",
        )
        self.outlook_summary.grid(row=3, column=0, padx=16, pady=12, sticky="ew")

        buttons = ctk.CTkFrame(tab)
        buttons.grid(row=4, column=0, padx=16, pady=10, sticky="ew")
        buttons.grid_columnconfigure((0, 1, 2), weight=1)
        ctk.CTkButton(buttons, text="Abrir lote atual", command=self.open_batch, height=44).grid(
            row=0, column=0, padx=7, pady=10, sticky="ew"
        )
        ctk.CTkButton(buttons, text="Marcar enviado e avançar", command=self.mark_sent, height=44).grid(
            row=0, column=1, padx=7, pady=10, sticky="ew"
        )
        ctk.CTkButton(buttons, text="Voltar um lote", command=self.go_back, height=44).grid(
            row=0, column=2, padx=7, pady=10, sticky="ew"
        )

        self.outlook_log = ctk.CTkTextbox(tab)
        self.outlook_log.grid(row=5, column=0, padx=16, pady=16, sticky="nsew")
        self._outlook_log("Os destinatários serão incluídos somente em CCO.")

    def select_todos(self) -> None:
        value = filedialog.askopenfilename(
            title="Selecione a planilha .Todos",
            filetypes=[("Planilhas Excel", "*.xlsx *.xlsm"), ("Todos os arquivos", "*.*")]
        )
        if value:
            self.todos_file = Path(value)
            self.todos_var.set(str(self.todos_file))

    def select_block(self) -> None:
        value = filedialog.askopenfilename(
            title="Selecione a planilha de bloqueados",
            filetypes=[("Planilhas Excel", "*.xlsx *.xlsm"), ("Todos os arquivos", "*.*")]
        )
        if value:
            self.block_file = Path(value)
            self.block_var.set(str(self.block_file))
            self.save_config()

    def select_output(self) -> None:
        value = filedialog.askdirectory(title="Selecione a pasta de saída")
        if value:
            self.output_folder = Path(value)
            self.output_var.set(str(self.output_folder))
            self.save_config()

    def process_async(self) -> None:
        if not self.todos_file or not self.todos_file.exists():
            messagebox.showwarning("Arquivo ausente", "Selecione a planilha .Todos.")
            return
        if not self.block_file or not self.block_file.exists():
            messagebox.showwarning("Arquivo ausente", "Selecione a planilha de bloqueados.")
            return
        try:
            batch_size = int(self.batch_var.get())
            if not 1 <= batch_size <= 600:
                raise ValueError
        except Exception:
            messagebox.showwarning("Valor inválido", "Informe um valor entre 1 e 600.")
            return

        self.busy.start()
        self.status_var.set("Limpando e comparando os e-mails...")
        threading.Thread(target=self._worker, daemon=True).start()

    def _worker(self) -> None:
        try:
            valid, invalid, raw_cells, candidate_count = ExcelReader.read_todos(self.todos_file)
            blocked_set, block_sheet = ExcelReader.read_blocklist(self.block_file)

            unique: list[str] = []
            duplicates: list[str] = []
            seen: set[str] = set()
            for email in valid:
                if email in seen:
                    duplicates.append(email)
                else:
                    seen.add(email)
                    unique.append(email)

            blocked = sorted(email for email in unique if email in blocked_set)
            final = [email for email in unique if email not in blocked_set]

            debate_removed: list[str] = []
            special_email = self.special_block_var.get().strip().lower()
            if self.type_var.get() == "Campanha especial" and special_email:
                corrected_special, _ = correct_email(special_email)
                if corrected_special and corrected_special in final:
                    final.remove(corrected_special)
                    debate_removed.append(corrected_special)

            result = ProcessResult(
                raw_cells=raw_cells,
                found_candidates=candidate_count,
                valid_before_dedup=len(valid),
                duplicates=duplicates,
                invalid=invalid,
                blocked=blocked,
                debate_removed=debate_removed,
                final=final,
                block_sheet=block_sheet,
            )

            size = int(self.batch_var.get())
            batches = [final[i:i + size] for i in range(0, len(final), size)]
            self.after(0, lambda: self.finish_processing(result, batches))
        except Exception as exc:
            log_error(traceback.format_exc())
            self.after(0, lambda: self.processing_error(exc))

    def finish_processing(self, result: ProcessResult, batches: list[list[str]]) -> None:
        self.busy.stop()
        self.result = result
        self.batches = batches
        self.current_batch = 0
        self.save_config()

        report = self.write_report()
        summary = self.build_summary()
        self.summary_label.configure(text=summary)
        self._result_log(summary.replace("\n", " | "))
        self._result_log(f"Relatório salvo em: {report}")
        self.refresh_outlook_summary()
        self.status_var.set("Lista pronta. Agora carregue o e-mail do Outlook.")

        messagebox.showinfo(
            "Processamento concluído",
            f"Lista final: {len(result.final):,} e-mails\n"
            f"Lotes: {len(batches)}\n\n"
            f"Relatório salvo em:\n{report}"
        )

    def processing_error(self, exc: Exception) -> None:
        self.busy.stop()
        self.status_var.set("Erro durante o processamento.")
        messagebox.showerror("Erro", str(exc))

    def campaign_folder(self) -> Path:
        root = self.output_folder or Path(self.output_var.get())
        root.mkdir(parents=True, exist_ok=True)
        name = safe_folder_name(self.outlook_subject or self.todos_file.stem if self.todos_file else "Divulgação")
        folder = root / name
        folder.mkdir(parents=True, exist_ok=True)
        return folder

    def write_report(self) -> Path:
        if self.result is None:
            raise RuntimeError("Nenhum resultado disponível.")
        if Workbook is None:
            raise RuntimeError("A biblioteca de Excel não foi incluída.")

        folder = self.campaign_folder()
        path = folder / "Relatorio_da_divulgacao.xlsx"
        workbook = Workbook()
        workbook.remove(workbook.active)

        green = PatternFill("solid", fgColor="1F6E4C")
        white_bold = Font(color="FFFFFF", bold=True)

        def add_sheet(name: str, headers: list[str], rows: list[list[object]]) -> None:
            sheet = workbook.create_sheet(name)
            sheet.append(headers)
            for row in rows:
                sheet.append(row)
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.fill = green
                cell.font = white_bold
            for column in sheet.columns:
                width = min(max(len(str(cell.value or "")) for cell in column) + 2, 70)
                sheet.column_dimensions[column[0].column_letter].width = width

        add_sheet("Resumo", ["Indicador", "Valor"], [
            ["Aplicativo", f"{APP_NAME} v{APP_VERSION}"],
            ["Desenvolvedor", "Deivid Brito"],
            ["Tipo", self.type_var.get()],
            ["Células lidas", self.result.raw_cells],
            ["Candidatos encontrados", self.result.found_candidates],
            ["Válidos antes dos repetidos", self.result.valid_before_dedup],
            ["Duplicados removidos", len(self.result.duplicates)],
            ["Inválidos descartados", len(self.result.invalid)],
            [f"Bloqueados ({self.result.block_sheet})", len(self.result.blocked)],
            ["Regra especial", len(self.result.debate_removed)],
            ["Total final", len(self.result.final)],
            ["Lotes", len(self.batches)],
            ["Data", datetime.now().strftime("%d/%m/%Y %H:%M:%S")],
        ])
        add_sheet("Lista Final", ["Email"], [[email] for email in self.result.final])
        add_sheet("Duplicados", ["Email"], [[email] for email in sorted(set(self.result.duplicates))])
        add_sheet("Inválidos", ["Conteúdo", "Motivo"], [[value, reason] for value, reason in self.result.invalid])
        add_sheet("Bloqueados", ["Email", "Motivo"], [
            [email, "Consta na planilha de bloqueados/retornos"] for email in self.result.blocked
        ])
        add_sheet("Regra especial", ["Email", "Motivo"], [
            [email, "Não enviar em campanhas especiais"] for email in self.result.debate_removed
        ])

        for index, batch in enumerate(self.batches, start=1):
            add_sheet(f"Lote {index}", ["Email"], [[email] for email in batch])

        workbook.save(path)
        (folder / "Lista_Final.txt").write_text("\n".join(self.result.final), encoding="utf-8")
        (folder / "Resumo_Final.txt").write_text(self.build_summary(), encoding="utf-8")
        return path

    def build_summary(self) -> str:
        if self.result is None:
            return "Nenhum processamento realizado."
        return (
            f"Tipo: {self.type_var.get()}\n"

            f"Válidos antes dos repetidos: {self.result.valid_before_dedup:,}\n"
            f"Duplicados removidos: {len(self.result.duplicates):,}\n"
            f"Inválidos descartados: {len(self.result.invalid):,}\n"
            f"Bloqueados removidos pela aba Resumo: {len(self.result.blocked):,}\n"
            f"Regra especial: {len(self.result.debate_removed):,}\n"
            f"TOTAL FINAL / ENVIADOS: {len(self.result.final):,}\n"
            f"Lotes: {len(self.batches):,}"
        )

    def load_outlook(self) -> None:
        try:
            info = OutlookBridge().selected_info()
            self.outlook_subject = info["subject"]
            self.outlook_var.set(
                f"Assunto limpo: {info['subject']}\n"
                f"Remetente: {info['sender']}\n"
                f"Recebido em: {info['received']}"
            )
            self._outlook_log(f"E-mail carregado: {info['subject']}")
            self.refresh_outlook_summary()
            if self.result:
                self.write_report()
            self.status_var.set("E-mail carregado. Abra o lote atual.")
        except Exception as exc:
            log_error(traceback.format_exc())
            messagebox.showerror("Outlook", str(exc))

    def open_batch(self) -> None:
        if not self.batches:
            messagebox.showwarning("Sem lotes", "Processe a lista primeiro.")
            return
        if self.current_batch >= len(self.batches):
            messagebox.showinfo("Concluído", "Todos os lotes já foram concluídos.")
            return
        try:
            recipients = self.batches[self.current_batch]
            subject = OutlookBridge().create_draft(recipients)
            self.outlook_subject = subject
            self._outlook_log(
                f"Lote {self.current_batch + 1}/{len(self.batches)} aberto "
                f"com {len(recipients)} destinatários em CCO."
            )
            self.status_var.set("Revise e envie no Outlook; depois marque o lote como enviado.")
        except Exception as exc:
            log_error(traceback.format_exc())
            messagebox.showerror("Outlook", str(exc))

    def mark_sent(self) -> None:
        if not self.batches:
            messagebox.showwarning("Sem lotes", "Processe a lista primeiro.")
            return
        if self.current_batch >= len(self.batches):
            messagebox.showinfo("Concluído", "Todos os lotes já foram concluídos.")
            return

        number = self.current_batch + 1
        self.current_batch += 1
        self._outlook_log(f"Lote {number} marcado como enviado.")
        self.refresh_outlook_summary()

        if self.current_batch >= len(self.batches):
            summary = self.build_summary()
            self.clipboard_clear()
            self.clipboard_append(summary)
            folder = self.campaign_folder()
            (folder / "Resumo_Final.txt").write_text(summary, encoding="utf-8")
            self.status_var.set("Divulgação concluída. Resumo copiado.")
            messagebox.showinfo(
                "DIVULGAÇÃO CONCLUÍDA",
                f"{summary}\n\nO resumo foi copiado para a área de transferência."
            )
        else:
            self.status_var.set(f"Próximo lote: {self.current_batch + 1} de {len(self.batches)}.")

    def go_back(self) -> None:
        if self.current_batch > 0:
            self.current_batch -= 1
            self.refresh_outlook_summary()
            self._outlook_log(f"Retornou para o lote {self.current_batch + 1}.")
        else:
            messagebox.showinfo("Primeiro lote", "Você já está no primeiro lote.")

    def refresh_outlook_summary(self) -> None:
        if not self.batches:
            self.outlook_summary.configure(text="Processe a lista antes de abrir os lotes.")
            return
        current = min(self.current_batch + 1, len(self.batches))
        self.outlook_summary.configure(
            text=(
                f"Total final: {len(self.result.final):,} • "
                f"Lote atual: {current} de {len(self.batches)} • "
                f"Assunto: {self.outlook_subject or 'ainda não carregado'}"
            )
        )

    def copy_summary(self) -> None:
        text = self.build_summary()
        self.clipboard_clear()
        self.clipboard_append(text)
        messagebox.showinfo("Resumo copiado", "O resumo foi copiado.")

    def export_again(self) -> None:
        try:
            path = self.write_report()
            messagebox.showinfo("Relatório salvo", str(path))
        except Exception as exc:
            messagebox.showerror("Erro", str(exc))

    def open_folder(self) -> None:
        try:
            os.startfile(self.campaign_folder())
        except Exception as exc:
            messagebox.showerror("Erro", str(exc))

    def save_config(self) -> None:
        save_json(CONFIG_FILE, {
            "block_file": str(self.block_file or ""),
            "output_folder": str(self.output_folder or self.output_var.get()),
            "batch_size": self.batch_var.get(),
            "special_block_email": self.special_block_var.get().strip(),
        })

    def _result_log(self, text: str) -> None:
        self.result_log.insert(END, f"[{datetime.now():%H:%M:%S}] {text}\n")
        self.result_log.see(END)

    def _outlook_log(self, text: str) -> None:
        self.outlook_log.insert(END, f"[{datetime.now():%H:%M:%S}] {text}\n")
        self.outlook_log.see(END)


if __name__ == "__main__":
    try:
        app = App()
        app.mainloop()
    except Exception:
        log_error(traceback.format_exc())
        raise
